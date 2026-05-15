#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""Extract review-critical materials from CSV/XLSX BOM files.

The script keeps the first loop deterministic: it parses a BOM, filters the
parts that deserve engineering attention, matches existing local device exports,
and emits bounded web lookup queries for evidence gathering.
"""

from __future__ import annotations

import argparse
import csv
import json
import re
import sys
import unicodedata
from dataclasses import dataclass
from pathlib import Path
from typing import Iterable


REPO_ROOT = Path(__file__).resolve().parent.parent
DEFAULT_EXPORT_DIR = REPO_ROOT / "data" / "sch_review_export"

SCHEMA_ID = "bom-key-materials/1.0"


COLUMN_ALIASES: dict[str, tuple[str, ...]] = {
    "item": (
        "item",
        "line",
        "no",
        "序号",
        "项次",
    ),
    "mpn": (
        "mpn",
        "manufacturer part number",
        "mfr part number",
        "mfg part number",
        "part number",
        "part no",
        "part #",
        "model",
        "device",
        "型号",
        "器件型号",
        "物料型号",
        "制造商料号",
        "厂商料号",
        "料号",
    ),
    "manufacturer": (
        "manufacturer",
        "mfr",
        "mfg",
        "vendor",
        "brand",
        "make",
        "厂商",
        "厂家",
        "制造商",
        "品牌",
    ),
    "designator": (
        "designator",
        "reference",
        "ref",
        "refdes",
        "refs",
        "reference designator",
        "位号",
        "元件位号",
    ),
    "quantity": (
        "quantity",
        "qty",
        "count",
        "数量",
        "用量",
    ),
    "value": (
        "value",
        "val",
        "nominal",
        "参数",
        "标称值",
        "规格",
    ),
    "package": (
        "package",
        "footprint",
        "case",
        "封装",
        "封装形式",
    ),
    "description": (
        "description",
        "desc",
        "comment",
        "part description",
        "notes",
        "备注",
        "描述",
        "说明",
        "名称",
        "物料描述",
    ),
    "category": (
        "category",
        "type",
        "class",
        "分类",
        "类别",
        "类型",
    ),
}


KEYWORD_RULES: tuple[tuple[str, str, tuple[str, ...]], ...] = (
    (
        "programmable_logic",
        "programmable logic or FPGA device",
        ("fpga", "cpld", "soc fpga", "programmable logic", "gowin", "lattice", "xilinx", "amd adaptive", "intel altera", "anlogic"),
    ),
    (
        "mcu_processor",
        "MCU/processor controls firmware, boot, clock, and debug behavior",
        ("mcu", "microcontroller", "microprocessor", "processor", "stm32", "gd32", "nrf", "esp32", "rp2040", "risc-v", "arm cortex"),
    ),
    (
        "power",
        "power IC affects rails, sequencing, thermal margin, or protection",
        (
            "pmic",
            "buck",
            "boost",
            "dc/dc",
            "dc-dc",
            "step-down",
            "step up",
            "step-up",
            "ldo",
            "regulator",
            "converter",
            "charger",
            "load switch",
            "efuse",
            "hot swap",
            "power mux",
            "ideal diode",
            "mosfet driver",
            "电源",
            "降压",
            "升压",
            "稳压",
            "转换器",
        ),
    ),
    (
        "clock",
        "clock source affects timing, jitter, startup, or interface compliance",
        ("oscillator", "crystal", "xo", "tcxo", "vcxo", "clock generator", "pll", "clk", "时钟", "晶振"),
    ),
    (
        "memory",
        "memory part affects boot, configuration, bandwidth, or retention",
        ("ddr", "sdram", "lpddr", "flash", "nor flash", "nand", "emmc", "eeprom", "fram", "sram", "memory"),
    ),
    (
        "high_speed_interface",
        "interface/transceiver can constrain signal integrity and protocol compatibility",
        (
            "ethernet phy",
            "phy",
            "serdes",
            "usb hub",
            "usb controller",
            "type-c",
            "type c",
            "pcie",
            "mipi",
            "lvds",
            "hdmi",
            "dp retimer",
            "retimer",
            "redriver",
            "transceiver",
            "rs485",
            "can transceiver",
        ),
    ),
    (
        "analog_mixed_signal",
        "analog/mixed-signal part needs operating range and application checks",
        ("adc", "dac", "op amp", "op-amp", "amplifier", "comparator", "mux", "multiplexer", "analog switch", "sensor"),
    ),
    (
        "connector",
        "connector affects mechanical fit, pinout, current rating, or interface compliance",
        ("connector", "socket", "header", "fpc", "ffc", "usb-c", "usb c", "sma connector", "hdmi receptacle", "连接器", "插座"),
    ),
    (
        "switch_control",
        "switch or user-control component can affect configuration and bring-up behavior",
        ("dip switch", "tact switch", "slide switch", "拨码开关", "自锁开关", "按键开关"),
    ),
    (
        "protection",
        "protection device affects ESD/surge/current fault behavior",
        ("tvs", "esd", "surge", "polyfuse", "fuse", "protection", "emi filter", "common mode choke"),
    ),
)


KEY_REF_PREFIXES = (
    "U",
    "IC",
    "FPGA",
    "CPU",
    "MCU",
    "PMIC",
    "ESD",
    "F",
    "FB",
    "D",
    "X",
    "Y",
    "OSC",
    "J",
    "P",
    "CN",
    "CON",
    "USB",
    "Q",
    "S",
    "T",
    "L",
)


PASSIVE_REF_PREFIXES = ("R", "C")
PASSIVE_KEYWORDS = (
    "resistor",
    "capacitor",
    "ceramic capacitor",
    "mlcc",
    "电阻",
    "电容",
)


MANUFACTURER_DOMAINS = {
    "texas instruments": "ti.com",
    "ti": "ti.com",
    "analog devices": "analog.com",
    "adi": "analog.com",
    "st": "st.com",
    "stmicroelectronics": "st.com",
    "microchip": "microchip.com",
    "nxp": "nxp.com",
    "onsemi": "onsemi.com",
    "infineon": "infineon.com",
    "qorvo": "qorvo.com",
    "lattice": "latticesemi.com",
    "gowin": "gowinsemi.com",
    "amd": "amd.com",
    "xilinx": "amd.com",
    "intel": "intel.com",
    "anlogic": "anlogic.com",
    "nexperia": "nexperia.com",
    "renesas": "renesas.com",
    "rohm": "rohm.com",
    "diodes": "diodes.com",
    "monolithic power systems": "monolithicpower.com",
    "mps": "monolithicpower.com",
    "molex": "molex.com",
    "murata": "murata.com",
    "tdk": "tdk.com",
    "abracon": "abracon.com",
    "samtec": "samtec.com",
    "bourns": "bourns.com",
    "yageo": "yageo.com",
    "do3think": "do3think.com",
}

KNOWN_MANUFACTURER_TOKENS = {
    "DO3THINK",
}


@dataclass(frozen=True)
class BomRow:
    line_number: int
    fields: dict[str, str]


def normalize_token(value: object) -> str:
    text = "" if value is None else str(value)
    text = unicodedata.normalize("NFKC", text)
    text = text.strip().lower()
    return re.sub(r"[\s_\-/#:：()（）.]+", "", text)


def normalize_text(value: object) -> str:
    text = "" if value is None else str(value)
    text = unicodedata.normalize("NFKC", text)
    text = text.replace("\u3000", " ")
    return re.sub(r"\s+", " ", text).strip()


def _alias_lookup() -> dict[str, str]:
    lookup: dict[str, str] = {}
    for canonical, aliases in COLUMN_ALIASES.items():
        lookup[normalize_token(canonical)] = canonical
        for alias in aliases:
            lookup[normalize_token(alias)] = canonical
    return lookup


ALIAS_LOOKUP = _alias_lookup()


def _canonical_column(header: object) -> str | None:
    return ALIAS_LOOKUP.get(normalize_token(header))


def _known_header_count(row: Iterable[object]) -> int:
    return sum(1 for cell in row if _canonical_column(cell))


def _find_header_row(rows: list[list[object]]) -> int:
    best_index = 0
    best_score = -1
    for index, row in enumerate(rows[:25]):
        score = _known_header_count(row)
        if score > best_score:
            best_score = score
            best_index = index
        if score >= 3:
            return index
    return best_index


def _rows_from_table(rows: list[list[object]]) -> list[BomRow]:
    rows = [row for row in rows if any(normalize_text(cell) for cell in row)]
    if not rows:
        return []

    header_index = _find_header_row(rows)
    header = rows[header_index]
    column_map: dict[int, str] = {}
    for index, cell in enumerate(header):
        canonical = _canonical_column(cell)
        if canonical:
            column_map[index] = canonical

    if not column_map:
        raise ValueError("Could not identify BOM header columns")

    bom_rows: list[BomRow] = []
    for offset, row in enumerate(rows[header_index + 1 :], start=header_index + 2):
        if _is_divider_row(row):
            continue
        fields = {name: "" for name in COLUMN_ALIASES}
        for index, canonical in column_map.items():
            if index < len(row):
                value = normalize_text(row[index])
                if value and not fields.get(canonical):
                    fields[canonical] = value
        if any(fields.values()):
            if _is_designator_continuation(fields) and bom_rows:
                current = bom_rows[-1].fields.get("designator", "")
                continuation = fields["designator"]
                separator = "," if current and not current.endswith(",") else ""
                bom_rows[-1].fields["designator"] = f"{current}{separator}{continuation}"
                continue
            bom_rows.append(BomRow(line_number=offset, fields=fields))
    return bom_rows


def _is_designator_continuation(fields: dict[str, str]) -> bool:
    return bool(fields.get("designator")) and not any(
        fields.get(name)
        for name in ("item", "quantity", "mpn", "manufacturer", "value", "package", "description", "category")
    )


def _is_divider_row(row: Iterable[object]) -> bool:
    text = "".join(normalize_text(cell) for cell in row)
    return bool(text) and bool(re.fullmatch(r"[_=\-\s]+", text))


def _read_csv(path: Path) -> list[list[str]]:
    raw = _read_text_with_fallback(path)
    sample = raw[:4096]
    if path.suffix.lower() in {".bom", ".tsv"} or sample.count("\t") > sample.count(","):
        return [list(row) for row in csv.reader(raw.splitlines(), dialect=csv.excel_tab)]
    try:
        dialect = csv.Sniffer().sniff(sample, delimiters=",;\t")
    except csv.Error:
        dialect = csv.excel_tab if path.suffix.lower() == ".tsv" else csv.excel
    return [list(row) for row in csv.reader(raw.splitlines(), dialect)]


def _read_text_with_fallback(path: Path) -> str:
    last_error: UnicodeDecodeError | None = None
    for encoding in ("utf-8-sig", "gb18030", "cp936", "latin-1"):
        try:
            return path.read_text(encoding=encoding)
        except UnicodeDecodeError as exc:
            last_error = exc
    if last_error:
        raise last_error
    return path.read_text()


def _read_xlsx(path: Path, sheet: str | None = None) -> list[list[object]]:
    try:
        import openpyxl
    except ImportError as exc:
        raise RuntimeError("XLSX BOM parsing requires openpyxl") from exc

    workbook = openpyxl.load_workbook(path, read_only=True, data_only=True)
    if sheet:
        worksheet = workbook[sheet]
    else:
        worksheet = workbook[workbook.sheetnames[0]]
    return [list(row) for row in worksheet.iter_rows(values_only=True)]


def load_bom(path: Path, sheet: str | None = None) -> list[BomRow]:
    suffix = path.suffix.lower()
    if suffix in {".xlsx", ".xlsm"}:
        rows = _read_xlsx(path, sheet=sheet)
    elif suffix in {".bom", ".csv", ".tsv", ".txt"}:
        rows = _read_csv(path)
    else:
        raise ValueError(f"Unsupported BOM file type: {path.suffix}")
    return _rows_from_table(rows)


def split_designators(value: str) -> list[str]:
    value = normalize_text(value)
    if not value:
        return []
    tokens = re.split(r"[,;，；\s]+", value)
    return [token.strip() for token in tokens if token.strip()]


def _designator_prefix(designator: str) -> str:
    match = re.match(r"([A-Za-z]+)", designator.strip())
    return match.group(1).upper() if match else ""


def _search_blob(fields: dict[str, str]) -> str:
    text = " ".join(
        fields.get(name, "")
        for name in ("mpn", "manufacturer", "description", "category", "value", "package")
    ).lower()
    return re.sub(r"[_\\/\-]+", " ", text)


def _contains_keyword(blob: str, keyword: str) -> bool:
    if re.fullmatch(r"[a-z0-9]+", keyword) and len(keyword) <= 3:
        return bool(re.search(rf"(?<![a-z0-9]){re.escape(keyword)}(?![a-z0-9])", blob))
    return keyword in blob or normalize_token(keyword) in normalize_token(blob)


def _passive_only(fields: dict[str, str], designators: list[str]) -> bool:
    if not designators:
        return False
    prefixes = {_designator_prefix(designator) for designator in designators}
    if not prefixes or not prefixes.issubset(PASSIVE_REF_PREFIXES):
        return False
    blob = _search_blob(fields)
    return any(keyword in blob for keyword in PASSIVE_KEYWORDS) or not fields.get("mpn")


def classify_key_material(fields: dict[str, str]) -> tuple[bool, str, list[str], str]:
    designators = split_designators(fields.get("designator", ""))
    if _passive_only(fields, designators):
        return False, "ordinary resistor/capacitor passive", [], "passive"

    blob = _search_blob(fields)
    matched_tags: list[str] = []
    reasons: list[str] = []
    for category, reason, keywords in KEYWORD_RULES:
        if any(_contains_keyword(blob, keyword) for keyword in keywords):
            matched_tags.append(category)
            reasons.append(reason)

    if matched_tags:
        return True, "; ".join(reasons[:2]), matched_tags, matched_tags[0]

    prefixes = {_designator_prefix(designator) for designator in designators}
    if prefixes.intersection(KEY_REF_PREFIXES):
        prefix = sorted(prefixes.intersection(KEY_REF_PREFIXES))[0]
        category = {
            "L": "power",
            "X": "clock",
            "Y": "clock",
            "OSC": "clock",
            "J": "connector",
            "P": "connector",
            "CN": "connector",
            "CON": "connector",
            "USB": "connector",
            "Q": "discrete_power",
            "S": "switch_control",
            "T": "magnetics",
            "ESD": "protection",
            "F": "protection",
            "FB": "protection",
            "D": "discrete_power",
        }.get(prefix, "integrated_circuit")
        return True, f"reference designator prefix {prefix} is review-critical", [category], category

    if fields.get("mpn") and fields.get("manufacturer") and fields.get("description"):
        return True, "identified manufacturer part with description needs evidence lookup", ["identified_part"], "identified_part"

    return False, "not enough evidence to treat as key material", [], "non_key"


def _looks_like_mpn(value: str) -> bool:
    token = normalize_text(value)
    if not token:
        return False
    if re.search(r"[\u4e00-\u9fff]", token):
        return False
    if re.fullmatch(r"\d{5,}", token):
        return True
    if any(ch.isdigit() for ch in token) and any(ch.isalpha() for ch in token):
        return True
    return False


def _is_electrical_spec(value: str) -> bool:
    token = normalize_text(value).lower()
    if token in {"", "nc", "dnp", "mark", "tp"}:
        return True
    return bool(
        re.fullmatch(
            r"\d+(\.\d+)?\s*(r|k|m|ohm|v|a|ma|uf|nf|pf|uh|nh|mhz|ppm|p|pin|x\d+p)",
            token,
        )
    )


def _is_package_token(value: str) -> bool:
    token = normalize_text(value).lower()
    package_patterns = (
        r"^(0[2468]02|1206|1210|1812)$",
        r"^(qfn|vqfn|bga|fbga|lqfp|tqfp|sop|sot|dfn|tqfn|tdfn|wson|hwson|udfn|x2son|son|mlf|rm|sc|sma|smc|pqfn).*$",
        r"^c(0[2468]02|1206|1210|1812)$",
        r".*\bpin$",
    )
    return any(re.match(pattern, token) for pattern in package_patterns)


def _looks_like_manufacturer(value: str) -> bool:
    token = normalize_text(value)
    if not token or len(token) > 40:
        return False
    if token.upper() in KNOWN_MANUFACTURER_TOKENS:
        return True
    if token.lower() in MANUFACTURER_DOMAINS:
        return True
    if any(ch.isdigit() for ch in token):
        return False
    return bool(re.fullmatch(r"[A-Za-z][A-Za-z .&+-]{1,30}", token))


def _infer_package(tokens: list[str]) -> str:
    package_patterns = (
        r"^(0[2468]02|1206|1210|1812)$",
        r"^(qfn|vqfn|bga|fbga|lqfp|tqfp|sop|sot|dfn|tqfn|tdfn|wson|hwson|udfn|x2son|son|mlf|rm|qfp|dip|bqb|type[- ]?c|qsfp).*$",
        r"^c(0[2468]02|1206|1210|1812)$",
    )
    for token in tokens:
        compact = token.strip().lower()
        if any(re.match(pattern, compact) for pattern in package_patterns):
            return token
    return ""


def _candidate_mpn_from_tokens(tokens: list[str], value: str) -> str:
    value = normalize_text(value)
    value_key = normalize_token(value)
    value_is_mpn = _looks_like_mpn(value) and not _is_electrical_spec(value) and not _is_package_token(value)
    candidates = [
        token
        for token in tokens
        if _looks_like_mpn(token)
        and not _is_electrical_spec(token)
        and not _is_package_token(token)
        and not _looks_like_manufacturer(token)
    ]

    if value_is_mpn:
        for candidate in candidates:
            candidate_key = normalize_token(candidate)
            if value_key and (value_key in candidate_key or candidate_key in value_key):
                return candidate
        return value

    if candidates:
        return candidates[-1]
    return ""


def enrich_bom_fields(fields: dict[str, str]) -> dict[str, str]:
    enriched = dict(fields)
    description = enriched.get("description", "")
    tokens = [normalize_text(token) for token in re.split(r"\\+", description) if normalize_text(token)]

    if tokens:
        if not enriched.get("manufacturer") and _looks_like_manufacturer(tokens[-1]):
            enriched["manufacturer"] = tokens[-1]
        if not enriched.get("mpn"):
            enriched["mpn"] = _candidate_mpn_from_tokens(tokens, enriched.get("value", ""))
        if not enriched.get("package"):
            enriched["package"] = _infer_package(tokens)

    if not enriched.get("description") and enriched.get("value"):
        enriched["description"] = enriched["value"]
    if not enriched.get("mpn") and _looks_like_mpn(enriched.get("value", "")) and not _is_electrical_spec(enriched.get("value", "")):
        enriched["mpn"] = enriched["value"]
    return enriched


def _part_key(value: str) -> str:
    return re.sub(r"[^A-Za-z0-9]+", "", value).upper()


def load_local_device_index(export_dir: Path = DEFAULT_EXPORT_DIR) -> dict[str, dict]:
    index: dict[str, dict] = {}
    if not export_dir.exists():
        return index

    for path in sorted(export_dir.glob("*.json")):
        if path.name.startswith("_"):
            continue
        try:
            payload = json.loads(path.read_text(encoding="utf-8"))
        except (OSError, json.JSONDecodeError):
            continue
        mpn = normalize_text(payload.get("mpn"))
        if not mpn:
            continue
        key = _part_key(mpn)
        index.setdefault(
            key,
            {
                "mpn": mpn,
                "manufacturer": normalize_text(payload.get("manufacturer")),
                "category": normalize_text(payload.get("category")),
                "description": normalize_text(payload.get("description")),
                "schema": payload.get("_schema"),
                "path": str(path.relative_to(REPO_ROOT)) if path.is_relative_to(REPO_ROOT) else str(path),
            },
        )
    return index


def local_match(fields: dict[str, str], index: dict[str, dict]) -> dict | None:
    mpn = fields.get("mpn", "")
    if not mpn:
        return None
    key = _part_key(mpn)
    if key in index:
        return index[key]

    # Some BOMs include orderable suffixes while local exports use a family MPN.
    candidates = []
    for local_key, record in index.items():
        if key.startswith(local_key) or local_key.startswith(key):
            candidates.append((len(local_key), record))
    if candidates:
        return sorted(candidates, key=lambda item: item[0], reverse=True)[0][1]
    return None


def manufacturer_domain(manufacturer: str) -> str | None:
    token = normalize_text(manufacturer).lower()
    return MANUFACTURER_DOMAINS.get(token)


def build_lookup_queries(fields: dict[str, str]) -> list[str]:
    mpn = fields.get("mpn", "")
    manufacturer = fields.get("manufacturer", "")
    description = fields.get("description", "")
    value = fields.get("value", "")
    package = fields.get("package", "")
    queries: list[str] = []

    if mpn and manufacturer:
        queries.append(f'"{mpn}" "{manufacturer}" datasheet official')
    if mpn:
        queries.append(f'"{mpn}" datasheet manufacturer product page')
    domain = manufacturer_domain(manufacturer)
    if mpn and domain:
        queries.append(f'site:{domain} "{mpn}" datasheet OR product')
    if mpn and package:
        queries.append(f'"{mpn}" "{package}" package pinout datasheet')
    if not mpn and description and manufacturer:
        queries.append(f'"{description}" "{manufacturer}" datasheet')
    elif not mpn and description:
        queries.append(f'"{description}" datasheet product page')
    if not mpn and value and value != description:
        queries.append(f'"{value}" datasheet product page')

    deduped: list[str] = []
    seen = set()
    for query in queries:
        if query and query not in seen:
            deduped.append(query)
            seen.add(query)
    return deduped[:4]


def extract_key_materials(bom_rows: list[BomRow], export_dir: Path = DEFAULT_EXPORT_DIR) -> dict:
    index = load_local_device_index(export_dir)
    key_materials = []
    skipped = 0

    for row in bom_rows:
        fields = enrich_bom_fields(row.fields)
        is_key, reason, risk_tags, category_guess = classify_key_material(fields)
        if not is_key:
            skipped += 1
            continue

        match = local_match(fields, index)
        evidence_status = "local_match" if match else "needs_web_lookup"
        key_materials.append(
            {
                "line_number": row.line_number,
                "designators": split_designators(fields.get("designator", "")),
                "quantity": fields.get("quantity", ""),
                "mpn": fields.get("mpn", ""),
                "manufacturer": fields.get("manufacturer", ""),
                "value": fields.get("value", ""),
                "package": fields.get("package", ""),
                "description": fields.get("description", ""),
                "category_guess": category_guess,
                "classification_reason": reason,
                "risk_tags": risk_tags,
                "local_match": match,
                "evidence_status": evidence_status,
                "lookup_queries": build_lookup_queries(fields) if not match else [],
            }
        )

    return {
        "_schema": SCHEMA_ID,
        "source": {
            "row_count": len(bom_rows),
        },
        "summary": {
            "total_rows": len(bom_rows),
            "key_material_count": len(key_materials),
            "skipped_row_count": skipped,
            "local_match_count": sum(1 for item in key_materials if item["local_match"]),
            "needs_web_lookup_count": sum(1 for item in key_materials if item["evidence_status"] == "needs_web_lookup"),
        },
        "key_materials": key_materials,
    }


def extract_key_materials_from_file(
    path: Path,
    *,
    sheet: str | None = None,
    export_dir: Path = DEFAULT_EXPORT_DIR,
) -> dict:
    bom_rows = load_bom(path, sheet=sheet)
    report = extract_key_materials(bom_rows, export_dir=export_dir)
    report["source"].update(
        {
            "path": str(path),
            "format": path.suffix.lower().lstrip("."),
        }
    )
    if sheet:
        report["source"]["sheet"] = sheet
    return report


def _print_summary(report: dict) -> None:
    summary = report["summary"]
    print(
        "BOM rows={total_rows} key_materials={key_material_count} local_matches={local_match_count} "
        "needs_web_lookup={needs_web_lookup_count}".format(**summary)
    )
    for item in report["key_materials"]:
        name = item["mpn"] or item["description"] or ",".join(item["designators"])
        print(
            f"- line {item['line_number']}: {name} [{item['category_guess']}] "
            f"{item['evidence_status']} - {item['classification_reason']}"
        )


def main(argv: list[str] | None = None) -> int:
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument("bom", type=Path, help="BOM file (.csv, .tsv, .txt, .xlsx, .xlsm)")
    parser.add_argument("--sheet", help="Worksheet name for XLSX/XLSM BOMs")
    parser.add_argument("--export-dir", type=Path, default=DEFAULT_EXPORT_DIR, help="Local sch-review export directory")
    parser.add_argument("--output", type=Path, help="Write JSON report to this path")
    parser.add_argument("--summary", action="store_true", help="Print a human-readable summary")
    args = parser.parse_args(argv)

    try:
        report = extract_key_materials_from_file(args.bom, sheet=args.sheet, export_dir=args.export_dir)
    except Exception as exc:
        print(f"ERROR: {exc}", file=sys.stderr)
        return 1

    rendered = json.dumps(report, indent=2, ensure_ascii=False)
    if args.output:
        args.output.write_text(rendered + "\n", encoding="utf-8")
    if args.summary:
        _print_summary(report)
    if not args.output and not args.summary:
        print(rendered)
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
