#!/usr/bin/env python3
"""Parse Gowin FPGA pinout XLSX files into structured JSON.

Input:  Gowin pinout .xlsx files (e.g. UG1110-1.0.4_GW5AR-25器件Pinout手册.xlsx)
Output: Multi-layer FPGA pin definition JSON (same schema as AMD parser)

Usage:
    python parse_gowin_pinout.py [xlsx_dir] [output_dir]
"""

import json
import re
import sys
from collections import defaultdict
from pathlib import Path

try:
    import openpyxl
except ImportError:
    print("pip install openpyxl")
    sys.exit(1)

DEFAULT_XLSX_DIR = Path(__file__).parent.parent / "data/raw/fpga/gowin/高云 FPGA"
DEFAULT_OUTPUT_DIR = Path(__file__).parent.parent / "data/extracted_v2/fpga/pinout"


def parse_gowin_xlsx(filepath: Path) -> list[dict]:
    """Parse a Gowin pinout XLSX. Returns list of results (one per package/sheet)."""
    wb = openpyxl.load_workbook(filepath, data_only=True)

    # Extract device name from filename: UG1110-1.0.4_GW5AR-25器件Pinout手册.xlsx → GW5AR-25
    fname = filepath.stem
    m = re.search(r"(GW5\w+-\d+)", fname)
    if not m:
        m = re.search(r"(AroraV)", fname)
    device = m.group(1) if m else fname

    # --- Parse Pin Definitions sheet (pin function descriptions) ---
    pin_defs = {}
    if "Pin Definitions" in wb.sheetnames:
        ws = wb["Pin Definitions"]
        for row in ws.iter_rows(min_row=2, values_only=True):
            if not row or not row[0]:
                continue
            name = str(row[0]).strip()
            direction = str(row[1]).strip() if row[1] else None
            desc = str(row[2]).strip() if len(row) > 2 and row[2] else None
            if name and direction:
                pin_defs[name.split("/")[0].split("[")[0].strip()] = {
                    "direction": direction,
                    "description": desc,
                }

    # --- Parse Power sheet ---
    power_rails = {}
    if "Power" in wb.sheetnames:
        ws = wb["Power"]
        for row in ws.iter_rows(min_row=1, values_only=True):
            if not row or not row[0]:
                continue
            name = str(row[0]).strip()
            desc = str(row[1]).strip() if len(row) > 1 and row[1] else ""
            # Find min/max voltage columns
            vmin = vmax = None
            for cell in row[2:]:
                if cell and "V" in str(cell):
                    val = str(cell).replace("V", "").strip()
                    try:
                        v = float(val)
                        if vmin is None:
                            vmin = v
                        else:
                            vmax = v
                    except ValueError:
                        pass
            if name.startswith("VCC") or name.startswith("VDD") or name.startswith("M0_") or name.startswith("VQPS"):
                power_rails[name] = {
                    "description": desc,
                    "min_voltage": vmin,
                    "max_voltage": vmax,
                }

    # --- Parse Pin List sheets (one per package) ---
    results = []
    for sheet_name in wb.sheetnames:
        if not sheet_name.startswith("Pin List"):
            continue

        ws = wb[sheet_name]
        # Extract package name from sheet name: "Pin List UG256P" → "UG256P"
        pkg_match = re.search(r"Pin List\s+(\S+)", sheet_name)
        package = pkg_match.group(1) if pkg_match else sheet_name.replace("Pin List", "").strip()
        if not package:
            package = "DEFAULT"

        # Parse header
        header = []
        for row in ws.iter_rows(min_row=1, max_row=1, values_only=True):
            header = [str(c).strip() if c else "" for c in row]
            break

        # Map column indices
        col_map = {}
        for i, h in enumerate(header):
            hl = h.lower() if h else ""
            if "管脚名称" in hl or "pin name" in hl or h == "管脚名称":
                col_map["name"] = i
            elif h == "功能" or (hl == "功能"):
                col_map["function"] = i
            elif h.upper() == "BANK" or h == "BANK":
                col_map["bank"] = i
            elif "dqs" in hl or h == "DQS":
                col_map["dqs"] = i
            elif "配置功能" in hl or h == "配置功能":
                col_map["config_func"] = i
            elif "差分" in hl or h == "差分Pair":
                col_map["diff_pair"] = i
            elif h == "LVDS" or hl == "lvds":
                col_map["lvds"] = i
            elif h == "X16" or hl == "x16":
                pass  # Skip X16 column
            elif package.upper() in h.upper():
                col_map["pin_loc"] = i

        # If pin_loc not found, use last non-mapped column as package pin
        if "pin_loc" not in col_map:
            for i in range(len(header) - 1, -1, -1):
                if i not in col_map.values() and header[i] and header[i] not in ("X16", ""):
                    col_map["pin_loc"] = i
                    # Also update package name from this header
                    if package == "DEFAULT" or not package:
                        package = header[i].strip()
                    break

        if "name" not in col_map or "pin_loc" not in col_map:
            continue

        # Parse pins
        pins = []
        for row in ws.iter_rows(min_row=2, values_only=True):
            if not row:
                continue
            name_val = row[col_map["name"]] if col_map["name"] < len(row) else None
            pin_loc = row[col_map["pin_loc"]] if col_map["pin_loc"] < len(row) else None

            if not name_val or not pin_loc:
                continue

            name_str = str(name_val).strip()
            pin_str = str(pin_loc).strip()
            if not name_str or not pin_str or pin_str == "None":
                continue

            func = str(row[col_map.get("function", -1)]).strip() if col_map.get("function") is not None and col_map["function"] < len(row) and row[col_map["function"]] else None
            bank = str(row[col_map.get("bank", -1)]).strip() if col_map.get("bank") is not None and col_map["bank"] < len(row) and row[col_map["bank"]] else None
            dqs = str(row[col_map.get("dqs", -1)]).strip() if col_map.get("dqs") is not None and col_map["dqs"] < len(row) and row[col_map["dqs"]] else None
            config_func = str(row[col_map.get("config_func", -1)]).strip() if col_map.get("config_func") is not None and col_map["config_func"] < len(row) and row[col_map["config_func"]] else None
            diff_pair = str(row[col_map.get("diff_pair", -1)]).strip() if col_map.get("diff_pair") is not None and col_map["diff_pair"] < len(row) and row[col_map["diff_pair"]] else None
            lvds = str(row[col_map.get("lvds", -1)]).strip() if col_map.get("lvds") is not None and col_map["lvds"] < len(row) and row[col_map["lvds"]] else None

            # Classify function
            classification = _classify_gowin_pin(name_str, func, bank, config_func)

            pin_entry = {
                "pin": pin_str,
                "name": name_str,
                "function": classification["function"],
                "bank": bank if bank and bank not in ("None", "none", "") else None,
                "config_function": config_func if config_func and config_func not in ("None", "none", "") else None,
                "dqs": dqs if dqs and dqs not in ("None", "none", "") else None,
                "lvds_capable": lvds == "True",
            }

            # Merge extra classification fields (polarity, serdes_quad, serdes_lane, etc.)
            for key in ("polarity", "serdes_quad", "serdes_lane"):
                if key in classification:
                    pin_entry[key] = classification[key]

            # Diff pair info from XLSX column (overrides classification polarity for IO pairs)
            if diff_pair and diff_pair not in ("None", "none", ""):
                if diff_pair.startswith("True_of_"):
                    pin_entry["polarity"] = "P"  # True = positive
                    pin_entry["diff_complement"] = diff_pair.replace("True_of_", "")
                elif diff_pair.startswith("Comp_of_"):
                    pin_entry["polarity"] = "N"  # Complement = negative
                    pin_entry["diff_complement"] = diff_pair.replace("Comp_of_", "")

            # DRC rules
            if classification.get("drc"):
                pin_entry["drc"] = classification["drc"]

            pins.append(pin_entry)

        if not pins:
            continue

        # --- Build diff pairs ---
        diff_pairs = _extract_gowin_diff_pairs(pins)

        # --- Build bank structure ---
        bank_structure = _build_gowin_banks(pins, power_rails)

        # --- Build DRC rules ---
        drc_rules = {
            "power_integrity": {"severity": "ERROR", "desc": "All power and ground pins must be connected"},
            "config_pins": {"severity": "ERROR", "desc": "Mandatory config pins (MODE, JTAGSEL, RECONFIG_N) must be connected"},
            "vcco_bank_consistency": {"severity": "ERROR", "desc": "All IO in same bank must use compatible IO standards sharing same VCCIO"},
            "diff_pair_integrity": {"severity": "ERROR", "desc": "Differential pairs must be used together or not at all"},
            "unused_io": {"severity": "WARNING", "desc": "Unused IO pins should not be left floating"},
            "jtag_pins": {"severity": "WARNING", "desc": "JTAG pins (TCK/TMS/TDI/TDO) should be properly connected or have pull resistors"},
        }

        # --- Build lookup ---
        lookup = {
            "pin_to_name": {p["pin"]: p["name"] for p in pins},
            "name_to_pin": {},
            "io_pins": [p["pin"] for p in pins if p["function"] == "IO"],
            "power_pins": [p["pin"] for p in pins if p["function"] in ("POWER", "GROUND")],
            "config_pins": [p["pin"] for p in pins if p["function"] == "CONFIG"],
        }
        # name_to_pin: only unique names
        name_counts = defaultdict(int)
        for p in pins:
            name_counts[p["name"]] += 1
        for p in pins:
            if name_counts[p["name"]] == 1:
                lookup["name_to_pin"][p["name"]] = p["pin"]

        # Summary
        func_counts = defaultdict(int)
        for p in pins:
            func_counts[p["function"]] += 1

        result = {
            "_schema_version": "2.0",
            "_purpose": "FPGA pin definition for LLM-driven schematic DRC",
            "_vendor": "Gowin",
            "device": device,
            "package": package,
            "source_file": filepath.name,
            "total_pins": len(pins),
            "summary": {
                "by_function": dict(sorted(func_counts.items())),
                "diff_pairs": {
                    "IO": len([p for p in diff_pairs if p["type"] == "IO"]),
                    "LVDS": len([p for p in diff_pairs if p["type"] == "LVDS"]),
                    "SERDES_RX": len([p for p in diff_pairs if p["type"] == "SERDES_RX"]),
                    "SERDES_TX": len([p for p in diff_pairs if p["type"] == "SERDES_TX"]),
                    "SERDES_REFCLK": len([p for p in diff_pairs if p["type"] == "SERDES_REFCLK"]),
                },
            },
            "power_rails": power_rails,
            "banks": bank_structure,
            "diff_pairs": diff_pairs,
            "drc_rules": drc_rules,
            "pins": pins,
            "lookup": lookup,
        }

        results.append(result)

    return results


def _classify_gowin_pin(name: str, func: str, bank: str, config_func: str) -> dict:
    """Classify a Gowin pin."""
    result = {"function": "OTHER"}

    name_upper = name.upper()
    func_str = (func or "").upper()

    # Check for config keywords in multi-function pin names (e.g. IOB175A/MODE0)
    config_keywords = ["TCK", "TMS", "TDI", "TDO", "MODE", "JTAGSEL", "RECONFIG_N",
                        "CCLK", "MOSI", "MISO", "SSPI_CS", "SSPI_CLK", "SSPI_WPN",
                        "DONE", "READY", "INIT", "EMCCLK", "D00", "D01", "D02", "D03",
                        "D04", "D05", "D06", "D07"]
    has_config = any(kw in name_upper for kw in config_keywords)

    if name_upper in ("GND", "VSS") or name_upper.startswith("GND") or name_upper.startswith("VSS"):
        result["function"] = "GROUND"
        result["drc"] = {"must_connect": True, "net": "GND"}
    elif "VCC" in name_upper or "VDD" in name_upper:
        result["function"] = "POWER"
        result["drc"] = {"must_connect": True}
    elif name_upper.startswith("M0_") or name_upper.startswith("M1_"):
        # MIPI pins
        result["function"] = "MIPI"
        if "CKP" in name_upper or "CKN" in name_upper:
            result["polarity"] = "P" if "CKP" in name_upper else "N"
        elif "P" in name_upper.split("_")[-1]:
            result["polarity"] = "P"
        elif "N" in name_upper.split("_")[-1]:
            result["polarity"] = "N"
    elif name_upper.startswith("ADC") or name_upper.startswith("ATEST"):
        result["function"] = "SPECIAL"
    elif re.match(r"Q\d+_LN\d+_(RX|TX)", name_upper):
        # SerDes transceiver pins (Q0_LN0_RXP_I, Q0_LN0_TXM_O, etc.)
        if "RX" in name_upper:
            result["function"] = "SERDES_RX"
        else:
            result["function"] = "SERDES_TX"
        if "RXP" in name_upper or "TXP" in name_upper:
            result["polarity"] = "P"
        elif "RXM" in name_upper or "TXM" in name_upper:
            result["polarity"] = "N"
        m_lane = re.search(r"Q(\d+)_LN(\d+)", name_upper)
        if m_lane:
            result["serdes_quad"] = int(m_lane.group(1))
            result["serdes_lane"] = int(m_lane.group(2))
    elif re.match(r"Q\d+_REF", name_upper):
        # SerDes reference clock
        result["function"] = "SERDES_REFCLK"
        m_q = re.search(r"Q(\d+)", name_upper)
        if m_q:
            result["serdes_quad"] = int(m_q.group(1))
        m_idx = re.search(r"_(\d+)$", name_upper)
        if m_idx:
            result["serdes_lane"] = int(m_idx.group(1))  # reuse lane for refclk index
        if "REFCLKP" in name_upper:
            result["polarity"] = "P"
        elif "REFCLKN" in name_upper or "REFCLKM" in name_upper:
            result["polarity"] = "N"
    elif name_upper == "NC" or name_upper.startswith("NC/"):
        result["function"] = "NC"
    elif has_config and (func_str == "I/O" or name_upper.startswith("IO")):
        # Multi-function IO with config capability
        result["function"] = "IO"
        # Identify the config function
        for kw in ["MODE0", "MODE1", "MODE2", "CCLK", "DONE", "READY", "RECONFIG_N",
                    "JTAGSEL_N", "TCK", "TMS", "TDI", "TDO", "MOSI", "MISO",
                    "SSPI_CS_N", "SSPI_CLK", "EMCCLK"]:
            if kw in name_upper:
                must = kw in ("MODE0", "MODE1", "MODE2", "CCLK", "DONE", "READY",
                              "RECONFIG_N", "MOSI", "SSPI_CS_N", "SSPI_CLK")
                result["drc"] = {
                    "config_function": kw,
                    "must_connect": must if must else "recommended",
                    "desc": f"Multi-function pin: IO + {kw}",
                }
                break
    elif func_str == "I/O" or name_upper.startswith("IO"):
        result["function"] = "IO"
    elif func_str in ("I", "O") and not name_upper.startswith("IO"):
        # Dedicated config pins
        result["function"] = "CONFIG"
        result["drc"] = {"must_connect": True, "desc": f"Dedicated: {name}"}

    return result


def _extract_gowin_diff_pairs(pins: list) -> list:
    """Extract differential pairs from Gowin pin data."""
    pairs = []
    true_pins = {}  # complement_name → pin
    comp_pins = {}  # complement_name → pin

    for p in pins:
        if p.get("polarity") == "P":
            comp_name = p.get("diff_complement", "")
            true_pins[comp_name] = p
        elif p.get("polarity") == "N":
            comp_name = p.get("diff_complement", "")
            comp_pins[comp_name] = p

    # Match: true_pins[X] has complement X, comp_pins[Y] has complement Y
    # true pin's complement = comp pin's name, comp pin's complement = true pin's name
    name_to_pin = {p["name"]: p for p in pins}

    for p in pins:
        if p.get("polarity") != "P":
            continue
        comp_name = p.get("diff_complement")
        if not comp_name or comp_name not in name_to_pin:
            continue
        comp_pin = name_to_pin[comp_name]
        if comp_pin.get("polarity") != "N":
            continue

        pair_type = "LVDS" if p.get("lvds_capable") else "IO"
        # Derive pair name from the True pin name
        base = re.sub(r"[AB]$", "", p["name"].split("/")[0])
        pairs.append({
            "type": pair_type,
            "pair_name": base,
            "p_pin": p["pin"],
            "n_pin": comp_pin["pin"],
            "p_name": p["name"],
            "n_name": comp_pin["name"],
            "bank": p.get("bank"),
        })

    # Also extract SerDes diff pairs
    serdes_by_key = defaultdict(dict)
    for p in pins:
        if p["function"] in ("SERDES_RX", "SERDES_TX", "SERDES_REFCLK"):
            # Key: function + quad + lane (or refclk index)
            quad = p.get("serdes_quad", "")
            lane = p.get("serdes_lane", "")
            pol = p.get("polarity")
            if p["function"] == "SERDES_REFCLK":
                key = f"REFCLK_Q{quad}_C{lane}"
            else:
                direction = "RX" if "RX" in p["function"] else "TX"
                key = f"{direction}_Q{quad}_L{lane}"
            if pol:
                serdes_by_key[key][pol] = p

    for key in sorted(serdes_by_key.keys()):
        entry = serdes_by_key[key]
        if "P" in entry and "N" in entry:
            pp = entry["P"]
            np = entry["N"]
            ptype = "SERDES_REFCLK" if "REFCLK" in key else ("SERDES_RX" if "RX" in key else "SERDES_TX")
            pairs.append({
                "type": ptype,
                "pair_name": key,
                "p_pin": pp["pin"],
                "n_pin": np["pin"],
                "p_name": pp["name"],
                "n_name": np["name"],
                "bank": pp.get("bank"),
            })

    return pairs


def _build_gowin_banks(pins: list, power_rails: dict) -> dict:
    """Build bank structure from pin data."""
    banks = defaultdict(lambda: {"io_type": None, "total_pins": 0, "io_pins": 0,
                                  "lvds_capable_count": 0, "clock_capable": []})

    for p in pins:
        if not p.get("bank"):
            continue
        b = banks[p["bank"]]
        b["total_pins"] += 1
        if p["function"] == "IO":
            b["io_pins"] += 1
            if p.get("lvds_capable"):
                b["lvds_capable_count"] += 1
            if p.get("config_function") and "GCLK" in p["config_function"]:
                b["clock_capable"].append(p["pin"])

    result = {}
    for bank_id in sorted(banks.keys(), key=lambda x: (not x.isdigit(), int(x) if x.isdigit() else 999)):
        b = banks[bank_id]
        entry = {
            "bank": bank_id,
            "total_pins": b["total_pins"],
            "io_pins": b["io_pins"],
            "lvds_capable_count": b["lvds_capable_count"],
            "clock_capable_pins": b["clock_capable"] if b["clock_capable"] else None,
        }

        # Match power rail
        vccio_key = f"VCCIO{bank_id}"
        if vccio_key in power_rails:
            rail = power_rails[vccio_key]
            entry["vccio_min"] = rail.get("min_voltage")
            entry["vccio_max"] = rail.get("max_voltage")
            entry["drc_note"] = f"VCCIO range: {rail.get('min_voltage')}V ~ {rail.get('max_voltage')}V"

        result[bank_id] = entry

    return result


def main():
    xlsx_dir = Path(sys.argv[1]) if len(sys.argv) > 1 else DEFAULT_XLSX_DIR
    output_dir = Path(sys.argv[2]) if len(sys.argv) > 2 else DEFAULT_OUTPUT_DIR
    output_dir.mkdir(parents=True, exist_ok=True)

    xlsx_files = sorted(xlsx_dir.glob("*.xlsx"))
    pinout_files = [f for f in xlsx_files if "Pinout" in f.name or "pinout" in f.name]

    if not pinout_files:
        print(f"No pinout XLSX files found in {xlsx_dir}")
        sys.exit(1)

    print(f"Found {len(pinout_files)} Gowin pinout files")

    total = 0
    for f in pinout_files:
        print(f"\nParsing {f.name}...")
        results = parse_gowin_xlsx(f)
        for result in results:
            safe_name = f"gowin_{result['device']}_{result['package']}".lower()
            out_path = output_dir / f"{safe_name}.json"
            with open(out_path, "w") as fp:
                json.dump(result, fp, indent=2, ensure_ascii=False)

            s = result["summary"]
            dp = s["diff_pairs"]
            print(f"  {result['device']} {result['package']}: {result['total_pins']} pins")
            print(f"  Functions: {json.dumps(s['by_function'])}")
            print(f"  Diff pairs: IO={dp.get('IO',0)} LVDS={dp.get('LVDS',0)}")
            print(f"  Banks: {len(result['banks'])}")
            print(f"  → {out_path}")
            total += 1

    print(f"\nTotal: {total} packages exported")


if __name__ == "__main__":
    main()
